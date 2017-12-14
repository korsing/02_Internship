import os
import xlsxwriter
import openpyxl
import urllib.request
import socket
from socket import timeout
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import ssl
import dns.resolver
import whois
import re
import sys

# 함수 정의 부분


# 0. 디렉토리를 재 설정
def changeDir(directory):
    #os.chdir("C:/Users/yjung/Desktop/Personal/Github/AKAMAI/CDN_Research/ImageCDN")
    os.chdir(directory)

# 1. CAT 리스트가 들어있는 원본 엑셀파일 열기
def open_excels(origin, destination, category):
    source = openpyxl.load_workbook(origin)
    sheet_load = source.get_sheet_by_name(category)
    destination = xlsxwriter.Workbook(destination)
    worksheet = destination.add_worksheet("Results")
    # 1-0. 저장 엑셀파일에 타이틀 작성
    writeExcel(worksheet, 0, 0, "Division")
    writeExcel(worksheet, 0, 1, "Website")
    writeExcel(worksheet, 0, 2, "CDN")
    writeExcel(worksheet, 0, 3, "Errors")

    # 1-1. url을 가져오기 위한 함수 호출
    get_domain(sheet_load, worksheet)

    # 1-3. 엑셀파일 저장 후 종료
    source.close()
    destination.close()


# 2. 원본 엑셀파일 sheet 이름을 받아와서 메인 도메인 url을 가져오고, 저장 엑셀파일에 기록하기
def get_domain(loadsheet, savesheet):
    counter = 1
    # 2-1. 1번째 행은 제목, 2번째 행부터 시작해서 끝까지 이동. (+1은 미만값 적용때문에 붙여줌)
    for i in range(1, loadsheet.max_row-1):
        if(loadsheet.cell(row=i+1, column=5).value == None and loadsheet.cell(row=i+2, column=5).value == None):
            maxrow = i
            break
    for rownum in range(2, maxrow + 1):  # 복사붙여넣기 용 : 9
        try:
            # 2-1-1. 로그
            print("현재 {}번째 웹사이트 방문 중입니다.".format(rownum - 1))
            original_url = loadsheet.cell(row=rownum, column=3).value
            # 2-1-2. 만약 원본 엑셀파일에 url이 존재하지 않는다면
            if (original_url == None):
                print("[에러] URL이 존재하지 않습니다.")
                print("")
                continue
            else:
                writeExcel(savesheet, counter, 0, "M")
                # 2-1-3. 만약 원본 엑셀파일에 url이 있다면, url 형식을 맞추기
                final_url = full_Url(original_url)
                writeExcel(savesheet, counter, 1, final_url)
                # 2-2. 메인 도메인에 CDN 검색, 그러기 위해서는 http를 빼야함.
                CDN_List = []
                cdn = find_Cname(urlparse(final_url)[1], savesheet, CDN_List)
                print("")
                content = ""
                if (len(cdn)):
                    for element in cdn:
                        if(element == None):
                            continue
                        else:
                            content += element + ", "
                    writeExcel(savesheet, counter, 2, content)
                else:
                    writeExcel(savesheet, counter, 2, "-")



                # 2-2-3. url에 방문해서 컨텐츠 crawl 해오기
                result = decode(final_url)
                # 2-2-4. 가지고 온 리스트 중에서 필요없는 내용들 다 버리기 (url이 아닌 것들, family site가 확실히 아닌것 등)
                blacklist = ["javascript",
                             "facebook", "twitter", "instagram", "youtube", "youtu.be",
                             "goo.gl", "google", "naver", "tistory", "cdn", "ftc",
                             "amazon", "aws", "amazonaws", "cloundflare", "rawgit",
                             "ad.about", "mailto", "linkedin", "slideshare", "github",
                             "flickr", "pinterest", "cloudfront", "wordpress", "cisco", "citrix", "netapp",
                             "hp.com", "microsoft", "apple", "vimeo", "surveymonkey",
                             "maxcdn", "jquery", "symantec", "norton", "cdnetworks", "kakao.", "ad.", "ad2."]
                # 2-4. 나온 결과값에서, 불필요한 내용들을 제거 (위 블랙리스트 참고)
                result = trimurl(result, blacklist)
                # 2-5. 나온 결과값에서 url 이쁘게 트리밍
                result = urlParse(result)
                # 2-6. 로그 : 결과값을 출력
                step = 1
                for i in range(len(result)):
                    if (urlparse(final_url)[1] != result[i]):
                        if("image" in result[i] or "img" in result[i]):
                            print("{:3} {}".format(step, result[i]))
                            writeExcel(savesheet, counter + 1, 0, "F")
                            writeExcel(savesheet, counter + 1, 1, result[i])
                            step += 1
                            # 2-6-1. 크롤된 웹사이트 CDN 검색
                            CDN_List = []
                            cdn = find_Cname(result[i], savesheet, CDN_List)
                            print("")
                            content = ""
                            if (len(cdn)):
                                for element in cdn:
                                    if(element == None):
                                        continue
                                    else:
                                        content += element + ", "
                                writeExcel(savesheet, counter + 1, 2, content)
                            else:
                                writeExcel(savesheet, counter + 1, 2, "-")
                            counter += 1

        except urllib.error.HTTPError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except socket.error:
            print("[에러] 소켓 통신에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except ssl.CertificateError:
            print("[에러] 인증서 인증에 실패하였습니다..")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except TimeoutError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except urllib.error.URLError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except urllib.error.HTTPError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except ConnectionError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except timeout:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")

        # 2.5. 한칸 띄워주고, 카운터변수 증가시켜주고 포문 종료
        print("")
        counter += 1


# 3. 원본 엑셀파일에서 가져온 url 형식 맞추기
def full_Url(url):
    # 3-1. 만약 url이 "http:"로 시작한다면 올바른 형식이니 그대로 유지
    if ("http:" in str(url)):
        full = url
    # 3-2. 만약 url이 "https:"로 시작한다면 올바른 형식이니 그대로 유지
    elif ("https:" in str(url)):
        full = url
    # 3-3. 나머지 경우에는 올바르지 않은 형식이니, 앞에 "http://"를 추가해줌
    else:
        full = "http://" + url
    # 3-4. 로그
    print(full, "홈페이지에 접근하고 있습니다.")
    return full


# 4. url를 타고 페이지에 접근해서 필요한 태그값을 crawl 해오기.
def decode(url):
    # 4-1. url 유효성을 검사
    request = urllib.request.Request(url)
    # 4-2. 5초안에 접속 못하면 탈락처리.
    data = urllib.request.urlopen(request, timeout=5).read()
    # 4-3. 웹사이트의 내용 전체를 긁어오기
    soup = BeautifulSoup(data, "lxml")
    # 4-4-1. "a" 태그에 "href"값을 가져오기
    link1 = [element.get("href") for element in soup.find_all("a")]
    # 4-4-2. "script" 태그에 "src"값을 가져오기
    link2 = [element.get("src") for element in soup.find_all("script")]
    # 4-4-3. "link" 태그에 "href"값을 가져오기
    link3 = [element.get("href") for element in soup.find_all("link")]
    # 4-4-4. "meta" 태그에 "content"값을 가져오기
    link4 = [element.get("content") for element in soup.find_all("meta")]
    # 4-4-5. "option" 태그에 "value"값을 가져오기
    link5 = [element.get("value") for element in soup.find_all("option")]
    # 4-4-6. "img" 태그에 "src"값을 가져오기
    link6 = [element.get("src") for element in soup.find_all("img")]
    # 4-4-7. "img" 태그에 "lowsrc"값을 가져오기
    link7 = [element.get("lowsrc") for element in soup.find_all("img")]
    # 4-4-8. "img" 태그에 "longdesc"값을 가져오기
    link8 = [element.get("longdesc") for element in soup.find_all("img")]
    # 4-5. 리스트들을 다 합쳐서 하나로 만들기
    result = link1 + link2 + link3 + link4 + link5 + link6 + link7 + link8
    return result


# 5. 불필요한 부분 제거
def trimurl(result, blacklist):
    # 5-1. 최종 결과를 저장할 리스트
    final = []
    # 5-2. 블랙리스트에 요소가 있다면 거짓
    flag = False
    for element in result:
        if (element == None):
            continue
        elif ("http:" in element or "https:" in element):
            for ban in blacklist:
                if (ban in element):
                    flag = False
                    break
                else:
                    flag = True
            if (flag == True):
                final.append(element)
        else:
            continue

    return final


# 6. 이쁘게 정리
def urlParse(result):
    # 6-1. 최종 결과를 저장할 리스트
    final = []
    for element in result:
        parse = urlparse(element)
        add = parse[1]
        if (add not in final):
            final.append(add)
    return final


# 7. CDN 종류
def find_Vendor(cDomain):
    category = {
        "Akamai": [".globalredir.akadns.net", ".akadns.net", ".edgekey.net", ".edgesuite.net", ".akamaiedge.net",".akamai.net", "aka"],
        "Level3": [".nstatic.net", ".footprint.net"],
        "Microsoft": [".azurewebsites.net", ".cloudapp.net", ".msecnd.net"],
        "Amazon": [".elb.amazonaws.com", ".amazonaws.com", ".cloudfront.net"],
        "Wix": [".wixdns.net", ".wix.com"],
        "OkServer": [".okserver.cn", "okserver.net"],
        "Cdnetworks": [".gccdn.net", ".cdnetworks.net", ".cdngc.net", ".cdngs.net", ".speedcdn.net", ".cdngl.net", ".cdnga.net", ".cdnetworks.com."],
        "Limelight Networks": [".llnwd.net", ".lldns.net"],
        "Verizon - EdgeCast": [".v0cdn.net", ".edgecastcdn.net", ".cedexis.net", ".mucdn.net", ".transactcdn.com.", "ecdns.net."],
        "Fastly": [".fastly.net"],
        "CloudFlare": [".cloudflare.net"],
        "Highwinds": [".hwcdn.net"],
        "CDN77(onApp)": [".cdn77.org"],
        "ChinaCache": [".lxsvc.cn", ".ccgslb.com", ".ccna.c3cdn.net"],
        "Yahoo": [".yahoo"],
        "Google": [".google.com", ".doubleclick.net", ".firebaseapp.com"],
        "Mirror Image": [".instacontent.net"],
        "Cachefly": [".cachefly.net"],
        "Coral Cache": [".nyucd.net"],
        "MaxCDN": [".netdna-cdn.com"],
        "NHN Corp.": [".navercdn.com", ".nheos.com"],
        "GS Neotek": [".gscdn.net", ".gscdn.com"],
        "LG": [".xcdn.com", ".cdn.cloudn.co.kr"],
        "KT": [".ktics.co.kr"],
        "Hyosung": [".hscdn.com"],
        "Penta Security Systems": [".cbricdns.com"],
        "NCIA Sheild Service": [".nciashield.org"],
        "Redhat": [".rhcloud.com"],
        "G-Core Labs S.A.": [".core.pw"],
        "SQUARESPACE": [".squarespace.com"],
        "KINX": [".kinxcdn.com"],
        "Whois Web": [".whoisweb.net"],
        "SKT Cloud" : [".c-cdn.tcloudbiz.com."],
        "Danal Infralab" : [".c-cdn.infralab.net.", ".gazelzone.com.", ".cdn.infralab.net."],
        "SK Broadband" : {".skcdn.co.kr.", ".skcdn.com", ".myskcdn.co.kr."},
        "CJ HelloVision" : {".cdn.visioncloud.co.kr."}
    }
    cdn_vendor = list(category.keys())

    for i in range(len(cdn_vendor)):
        for element in category.get(cdn_vendor[i]):
            if (element in cDomain):
                return cdn_vendor[i]


# 8. 메인 홈페이지 CDN 사용 여부
def find_Cname(domain, savesheet, CDN_List):
    # try:
    my_resolver = dns.resolver.Resolver()
    my_resolver.nameservers = ['8.8.8.8']
    CDN_List = []
    try:
        while (1):
            for rdata in dns.resolver.query(domain.strip(), "CNAME"):
                cDomain = str(rdata)

                if(find_Vendor(cDomain) == None):
                    if("cdn" in cDomain):
                        print("\t\tCDN Vendor : Unknown")
                        CDN_List.append("Unknown")
                    else:
                        continue
                else:
                    print("\t\tCDN Vendor : ", find_Vendor(cDomain))
                    CDN_List.append(find_Vendor(cDomain))
                # print("2 : ", CDN_List)
                # print("3 : ", cDomain)
            domain = cDomain[:len(cDomain) - 1]
        CDN_List = set(CDN_List)
        # print(CDN_List)
        return CDN_List
    except:
        try:
            for rdata in dns.resolver.query(domain.strip(), "A"):

                if(len(CDN_List) == 0):
                    return []
                else:
                    # print("\t\tARECORD :", rdata.address)
                    CDN_List = set(CDN_List)
                    # print(CDN_List)
                    return CDN_List
        except:
            print("No CNAME or ARECORD found.")
            return []




# 11. 엑셀에 작성
def writeExcel(file, row, column, content):
    file.write(row, column, content)


# 메인 함수
if __name__ == "__main__":
        print("Akamai Korea")
        if (len(sys.argv) < 3):
            print("CDN Research Finder (Ver 1.0) has been launched without any settings..")
            print("Please execute the program with according details..")
            print("Example : python CND_Research.py Akamai_CAT_List.xlsx CAT-4")
        else:
            origin = sys.argv[1]
            destination = "CDN_Research(" + str(sys.argv[2]) + ").xlsx"
            category = sys.argv[2]

            # 0. 디렉토리 변경
            while(1):
                directory = input("Please enter the current location : ")
                changeDir(directory)
                # 1. 원본 엑셀파일 열기
                try:
                    open_excels(origin, destination, category)
                    print("프로그램이 정상적으로 종료되었습니다.")
                    break
                except OSError:
                    print("Illegal directory has been entered. Please enter a correct directory where the program is saved.")










# 메인 함수
