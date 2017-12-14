import os
import xlsxwriter
import openpyxl
import urllib.request
import socket
from socket import timeout
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import ssl
import dns.resolver
import whois
import re

# 함수 정의 부분


# 0. 디렉토리를 재 설정
def changeDir():
    os.chdir("C:/Users/yjung/Desktop/Personal/Github/AKAMAI/CDN_Research")

# 1. CAT 리스트가 들어있는 원본 엑셀파일 열기
def open_excels():
    source = openpyxl.load_workbook("CAT_List_Akamai.xlsx")
    sheet_load = source.get_sheet_by_name("CAT-3")
    destination = xlsxwriter.Workbook("CDN_Research(CAT-3).xlsx")
    worksheet = destination.add_worksheet("Results")
    # 1-0. 저장 엑셀파일에 타이틀 작성
    writeExcel(worksheet, 0, 0, "Division")
    writeExcel(worksheet, 0, 1, "Website")
    writeExcel(worksheet, 0, 2, "CDN")
    writeExcel(worksheet, 0, 3, "Errors")
    writeExcel(worksheet, 0, 4, "Name")
    writeExcel(worksheet, 0, 5, "Phone")
    writeExcel(worksheet, 0, 6, "Email")
    # 1-1. url을 가져오기 위한 함수 호출
    get_domain(sheet_load, worksheet)
    
    # 1-3. 엑셀파일 저장 후 종료
    source.close()
    destination.close()

# 2. 원본 엑셀파일 sheet 이름을 받아와서 메인 도메인 url을 가져오고, 저장 엑셀파일에 기록하기
def get_domain(loadsheet, savesheet):
    counter = 1
    # 2-1. 1번째 행은 제목, 2번째 행부터 시작해서 끝까지 이동. (+1은 미만값 적용때문에 붙여줌)
    for rownum in range(2, 184):  # 복사붙여넣기 용 : loadsheet.max_row+1
        try:
            # 2-1-1. 로그
            print("현재 {}번째 웹사이트 방문 중입니다.".format(rownum-1))
            original_url = loadsheet.cell(row=rownum, column=3).value
            # 2-1-2. 만약 원본 엑셀파일에 url이 존재하지 않는다면
            if (original_url == None):
                print("[에러] URL이 존재하지 않습니다.")
                writeExcel(savesheet, counter, 0, "M")
                writeExcel(savesheet, counter, 1, "-")
                writeExcel(savesheet, counter, 2, "-")
                writeExcel(savesheet, counter, 3, "- URL 없음 -")
                writeExcel(savesheet, counter, 4, "-")
                writeExcel(savesheet, counter, 5, "-")
                writeExcel(savesheet, counter, 6, "-")
                counter += 1
            else:
                writeExcel(savesheet, counter, 0, "M")
            # 2-1-3. 만약 원본 엑셀파일에 url이 있다면, url 형식을 맞추기
                final_url = full_Url(original_url)
            # 2-2. 메인 도메인에 CDN 검색, 그러기 위해서는 http를 빼야함.
                CDN_List = []
                cdn = find_Cname(urlparse(final_url)[1], savesheet, CDN_List)
                print("")
                content = ""
                if(len(cdn)):
                    for element in cdn:
                        content += element + ","
                    writeExcel(savesheet, counter, 2, content)
                else:
                    writeExcel(savesheet, counter, 2, "-")
            # 2-3. 메인 도메인의 후이즈 정보 출력
            regex = re.compile(r"https?://www\.?")
            url = regex.sub("", final_url)
            info = get_Whois(url)
            if(info != 0):
                # 2-3-1. 이름 정보 출력
                if ("name" in info.keys()):
                    writeExcel(savesheet, counter, 4, str(info["name"]))
                elif ("admin_name" in info.keys()):
                    writeExcel(savesheet, counter, 4, str(info["admin_name"]))
                else:
                    writeExcel(savesheet, counter, 4, "-")
                # 2-3-2. 이메일 정보 출력
                if ("email" in info.keys()):
                    total = trimWhois("email", info)
                    # print("email : ", type(total))
                    writeExcel(savesheet, counter, 6, total)
                elif ("emails" in info.keys()):
                    total = trimWhois("emails", info)
                    # print("emails : ", type(total))
                    writeExcel(savesheet, counter, 6, total)
                elif ("admin_email" in info.keys()):
                    total = trimWhois("admin_email", info)
                    # print("admin_email : ", type(total))
                    writeExcel(savesheet, counter, 6, total)
                else:
                    writeExcel(savesheet, counter, 6, "-")
                # 2-3-3. 전화번호 정보 출력
                if("admin_phone" in info.keys()):
                    total = trimWhois("admin_phone", info)
                    # print("admin_phone : ", type(total))
                    writeExcel(savesheet, counter, 5, total)
            else:
                print("해당 도메인에 대한 Whois 정보가 없습니다.")
                writeExcel(savesheet, counter, 4, "-")
                writeExcel(savesheet, counter, 5, "-")
                writeExcel(savesheet, counter, 6, "-")

            writeExcel(savesheet, counter, 1, final_url)

            # 2-2-3. url에 방문해서 컨텐츠 crawl 해오기
            result = decode(final_url)
            # 2-2-4. 가지고 온 리스트 중에서 필요없는 내용들 다 버리기 (url이 아닌 것들, family site가 확실히 아닌것 등)
            blacklist = ["javascript",
                         "facebook", "twitter", "instagram", "youtube", "youtu.be",
                         "goo.gl", "google", "naver", "tistory", "cdn", "ftc",
                         "amazon", "aws", "amazonaws", "cloundflare", "rawgit",
                         "ad.about", "mailto", "linkedin", "slideshare", "github",
                         "flickr", "pinterest", "cloudfront", "wordpress", "cisco", "citrix", "netapp",
                         "hp.com", "microsoft", "apple", "vimeo", "surveymonkey",
                         "maxcdn", "jquery", "symantec", "norton", "cdnetworks", "kakao.", "ad.", "ad2."]
            # 2-4. 나온 결과값에서, 불필요한 내용들을 제거 (위 블랙리스트 참고)
            result = trimurl(result, blacklist)
            # 2-5. 나온 결과값에서 url 이쁘게 트리밍
            result = urlParse(result)
            # 2-6. 로그 : 결과값을 출력
            step = 1
            for i in range(len(result)):
                if(urlparse(final_url)[1] != result[i]):
                    print("{:3} {}".format(step, result[i]))
                    writeExcel(savesheet, counter+1, 0, "F")
                    writeExcel(savesheet, counter+1, 1, result[i])
                    step += 1
                    # 2-6-1. 크롤된 웹사이트 CDN 검색
                    CDN_List = []
                    cdn = find_Cname(result[i], savesheet, CDN_List)
                    print("")
                    content = ""
                    if (len(cdn)):
                        for element in cdn:
                            content += element + ","
                        writeExcel(savesheet, counter+1, 2, content)
                    else:
                        writeExcel(savesheet, counter+1, 2, "-")
            # 2-7. 크롤된 도메인에 대한 Whois 정보 검색
                    info = get_Whois(result[i])
                    print(info)
                    if (info != 0):
                        # 2-3-1. 이름 정보 출력
                        if ("name" in info.keys()):
                            writeExcel(savesheet, counter+1, 4, str(info["name"]))
                        elif ("admin_name" in info.keys()):
                            writeExcel(savesheet, counter+1, 4, str(info["admin_name"]))
                        else:
                            writeExcel(savesheet, counter+1, 4, "-")
                        # 2-3-2. 이메일 정보 출력
                        if ("email" in info.keys()):
                            total = trimWhois("email", info)
                            # print("email : ", type(total))
                            writeExcel(savesheet, counter+1, 6, total)
                        elif ("emails" in info.keys()):
                            total = trimWhois("emails", info)
                            # print("emails : ", type(total))
                            writeExcel(savesheet, counter+1, 6, total)
                        elif ("admin_email" in info.keys()):
                            total = trimWhois("admin_email", info)
                            # print("admin_email : ", type(total))
                            writeExcel(savesheet, counter+1, 6, total)
                        else:
                            writeExcel(savesheet, counter+1, 6, "-")
                        # 2-3-3. 전화번호 정보 출력
                        if ("admin_phone" in info.keys()):
                            total = trimWhois("admin_phone", info)
                            # print("admin_phone : ", type(total))
                            writeExcel(savesheet, counter+1, 5, total)
                    else:
                        print("[에러] 해당 도메인에 대한 Whois 정보가 없습니다.")
                        writeExcel(savesheet, counter+1, 4, "-")
                        writeExcel(savesheet, counter+1, 5, "-")
                        writeExcel(savesheet, counter+1, 6, "-")
                    counter += 1

        except urllib.error.HTTPError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except socket.error:
            print("[에러] 소켓 통신에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except ssl.CertificateError:
            print("[에러] 인증서 인증에 실패하였습니다..")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except TimeoutError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except urllib.error.URLError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except urllib.error.HTTPError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except ConnectionError:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")
        except timeout:
            print("[에러] 접속에 실패하였습니다.")
            writeExcel(savesheet, counter, 3, "사이트 접속 불가")

        # 2.5. 한칸 띄워주고, 카운터변수 증가시켜주고 포문 종료
        print("")
        counter += 1

# 3. 원본 엑셀파일에서 가져온 url 형식 맞추기
def full_Url(url):
    # 3-1. 만약 url이 "http:"로 시작한다면 올바른 형식이니 그대로 유지
    if ("http:" in str(url)):
        full = url
    # 3-2. 만약 url이 "https:"로 시작한다면 올바른 형식이니 그대로 유지
    elif ("https:" in str(url)):
        full = url
    # 3-3. 나머지 경우에는 올바르지 않은 형식이니, 앞에 "http://"를 추가해줌
    else:
        full = "http://" + url
    # 3-4. 로그
    print(full, "홈페이지에 접근하고 있습니다.")
    return full

# 4. url를 타고 페이지에 접근해서 필요한 태그값을 crawl 해오기.
def decode(url):
    # 4-1. url 유효성을 검사
    request = urllib.request.Request(url)
    # 4-2. 5초안에 접속 못하면 탈락처리.
    data = urllib.request.urlopen(request, timeout=5).read()
    # 4-3. 웹사이트의 내용 전체를 긁어오기
    soup = BeautifulSoup(data, "lxml")
    # 4-4-1. "a" 태그에 "href"값을 가져오기
    link1 = [element.get("href") for element in soup.find_all("a")]
    # 4-4-2. "script" 태그에 "src"값을 가져오기
    link2 = [element.get("src") for element in soup.find_all("script")]
    # 4-4-3. "link" 태그에 "href"값을 가져오기
    link3 = [element.get("href") for element in soup.find_all("link")]
    # 4-4-4. "meta" 태그에 "content"값을 가져오기
    link4 = [element.get("content") for element in soup.find_all("meta")]
    # 4-4-5. "option" 태그에 "value"값을 가져오기
    link5 = [element.get("value") for element in soup.find_all("option")]
    # 4-5. 리스트들을 다 합쳐서 하나로 만들기
    result = link1 + link2 + link3 + link4 + link5
    return result

# 5. 불필요한 부분 제거
def trimurl(result, blacklist):
    # 5-1. 최종 결과를 저장할 리스트
    final = []
    # 5-2. 블랙리스트에 요소가 있다면 거짓
    flag = False
    for element in result:
        if(element == None):
            continue
        elif("http:" in element or "https:" in element):
            for ban in blacklist:
                if(ban in element):
                    flag = False
                    break
                else:
                    flag = True
            if(flag == True):
                final.append(element)
        else:
            continue

    return final

# 6. 이쁘게 정리
def urlParse(result):
    # 6-1. 최종 결과를 저장할 리스트
    final = []
    for element in result:
        parse = urlparse(element)
        add = parse[1]
        if(add not in final):
            final.append(add)
    return final

# 7. CDN 종류
def find_Vendor(cDomain):
    category = {
        "Akamai":                         [".globalredir.akadns.net", ".akadns.net", ".edgekey.net", ".edgesuite.net", ".akamaiedge.net", ".akamai.net"],
        "Level3":                         [".nstatic.net", ".footprint.net"],
        "Microsoft":                      [".azurewebsites.net", ".cloudapp.net", ".msecnd.net"],
        "Amazon":                         [".elb.amazonaws.com", ".amazonaws.com", ".cloudfront.net"],
        "Wix":                            [".wixdns.net", ".wix.com"],
        "OkServer Internet Solution Inc": [".okserver.cn", "okserver.net"],
        "Cdnetworks":                     [".gccdn.net", ".cdnetworks.net", ".cdngc.net", ".cdngs.net", ".speedcdn.net", ".cdngl.net",".cdnga.net"],
        "Limelight Networks":             [".llnwd.net", ".lldns.net"],
        "Verizon - EdgeCast":             [".v0cdn.net", ".edgecastcdn.net", ".cedexis.net", ".mucdn.net"],
        "Fastly":                         [".fastly.net"],
        "CloudFlare":                     [".cloudflare.net"],
        "Highwinds":                      [".hwcdn.net"],
        "CDN77(onApp)":                   [".cdn77.org"],
        "ChinaCache":                     [".lxsvc.cn", ".ccgslb.com", ".ccna.c3cdn.net"],
        "Yahoo":                          [".yahoo"],
        "Google":                         [".google.com", ".doubleclick.net", ".firebaseapp.com"],
        "Mirror Image":                   [".instacontent.net"],
        "Cachefly":                       [".cachefly.net"],
        "Coral Cache":                    [".nyucd.net"],
        "MaxCDN":                         [".netdna-cdn.com"],
        "NHN Corp.":                      [".navercdn.com", ".nheos.com"],
        "GS Neotek":                      [".gscdn.net", ".gscdn.com"],
        "LG":                             [".xcdn.com", ".cdn.cloudn.co.kr"],
        "KT":                             [".ktics.co.kr"],
        "Hyosung":                        [".hscdn.com"],
        "Penta Security Systems":         [".cbricdns.com"],
        "NCIA Sheild Service":            [".nciashield.org"],
        "Redhat":                         [".rhcloud.com"],
        "G-Core Labs S.A.":               [".core.pw"],
        "SQUARESPACE":                    [".squarespace.com"],
        "KINX":                           [".kinxcdn.com"],
        "Whois Web":                      [".whoisweb.net"]
    }
    cdn_vendor = list(category.keys())

    for i in range(len(cdn_vendor)):
        for element in category.get(cdn_vendor[i]):
            if (element in cDomain):
                return cdn_vendor[i]


# 8. 메인 홈페이지 CDN 사용 여부
def find_Cname(domain, savesheet, CDN_List):
    try:
        for rdata in dns.resolver.query(domain.strip(), "CNAME"):
            cDomain = str(rdata)
            print("    [CNAME 발견] {:10} : {}".format(find_Vendor(cDomain), cDomain))
            CDN_List.append(find_Vendor(cDomain))
            find_Cname(cDomain[0:len(cDomain) - 1])
    except:
        find_ARecord(domain, savesheet)
    return CDN_List

def find_ARecord(domain, savesheet):
    try:
        for rdata in dns.resolver.query(domain, "A"):
            print("    [ARECORD 발견] : {}".format(rdata.address))
    except:
        pass

#9. 후이즈 정보 출력
def get_Whois(domain):
    try:
        print("{} 후이즈 확인중".format(domain))
        info = whois.whois(domain)
        print("{} 후이즈 확인 완료".format(domain))
        return info
    except Exception:
        print("{} 후이즈 에러".format(domain))
        return 0

# 10. 후이즈 정보 정리
def trimWhois(key, dictionary):
    total = []
    if (type(dictionary[key]) == list):
        total = ""
        for element in dictionary[key]:
            total += element + ", "
        return total
    elif (type(dictionary[key]) == str):
        total = dictionary[key]
        return total
    else:
        return "-"


# 11. 엑셀에 작성
def writeExcel(file, row, column, content):
    file.write(row, column, content)


    
# 메인 함수
if __name__ == "__main__":
    # 0. 디렉토리 변경
    changeDir()
    # 1. 원본 엑셀파일 열기
    open_excels()



    
    
    
    
    
    
    
# 메인 함수
