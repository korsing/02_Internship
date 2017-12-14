from openpyxl import load_workbook
import dns.resolver
import sys
from colors import *

wb = load_workbook('cat.xlsx')
ws = wb['All']

def find_Vendor(cDomain):
  #Akamai
    if cDomain.find(".globalredir.akadns.net") > -1:
      return ("Akamai - ChinaCDN")
  
    if cDomain.find(".akadns.net") > -1:
      return ("Akamai - GTM")
  
    if cDomain.find(".edgekey.net") > -1:
      return ("Akamai - ESSL")
  
    if cDomain.find(".edgesuite.net") > -1:
      return ("Akamai - FF")
  
    if cDomain.find(".akamaiedge.net") > -1:
      return ("Akamai - ESSL")
  
    if cDomain.find(".akamai.net") > -1:
      return ("Akamai - FF")

    #Other CDN 
    if cDomain.find(".nsatc.net") > -1:
      return ("Level 3 Communications")
  
    if cDomain.find(".footprint.net") > -1:
      return ("Level3")

    if cDomain.find(".azurewebsites.net") > -1:
      return ("Microsft Azure")

    if cDomain.find(".cloudapp.net") > -1:
      return ("Microsft Azure")
  
    if cDomain.find(".msecnd.net") > -1:
      return ("Microsoft Distribution")

    if cDomain.find(".elb.amazonaws.com") > -1:
      return ("Amazon Web Services - ELB")
        
    if cDomain.find(".amazonaws.com") > -1:
      return ("Amazon Web Services")
  
    if cDomain.find(".cloudfront.net") > -1:
      return ("Amazon Web Services - CloudFront")

    #
    if cDomain.find(".wixdns.net") > -1:
      return ("Wix Web Site Templates")

    if cDomain.find(".wix.com") > -1:
      return ("Wix Web Site Templates")

    #
    if cDomain.find(".okserver.cn") > -1:
      return ("OkServer Internet Solution Inc")

    if cDomain.find(".okserver.net") > -1:
      return ("OkServer Internet Solution Inc")

    #
    if cDomain.find(".gccdn.net") > -1:
      return ("CDNETWORKS")
  
    if cDomain.find(".cdnetworks.net") > -1:
      return ("CDNETWORKS")
    
    if cDomain.find(".cdngc.net") > -1:
      return ("CDNETWORKS")
  
    if cDomain.find(".cdngs.net") > -1:
      return ("CDNETWORKS")
  
    if cDomain.find(".speedcdn.net") > -1:
      return ("CDNETWORKS")

    if cDomain.find(".cdngl.net") > -1:
      return ("CDNETWORKS")

    if cDomain.find(".cdnga.net") > -1:
      return ("CDNETWORKS")

    if cDomain.find(".llnwd.net") > -1:
      return ("Limelight Networks")
  
    if cDomain.find(".lldns.net") > -1:
      return ("Limelight Networks")

    if cDomain.find(".v0cdn.net") > -1:
      return ("Verizon - EdgeCast")
  
    if cDomain.find(".edgecastcdn.net") > -1:
      return ("Verizon - EdgeCast")
  
    if cDomain.find(".cedexis.net") > -1:
      return ("Verizon - EdgeCast")
  
    if cDomain.find(".mucdn.net") > -1:
      return ("Verizon - EdgeCast")

    if cDomain.find(".fastly.net") > -1:
      return ("Fastly")

    if cDomain.find(".cloudflare.net") > -1:
      return ("CloudFlare")

    if cDomain.find(".hwcdn.net") > -1:
      return ("Highwinds")

    if cDomain.find(".cdn77.org") > -1:
      return ("CDN77(onApp)")

    if cDomain.find(".lxsvc.cn") > -1:
      return ("ChinaCache")
  
    if cDomain.find(".ccgslb.com") > -1:
      return ("ChinaCache")
  
    if cDomain.find(".ccna.c3cdn.net") > -1:
      return ("ChinaCache")

    if cDomain.find(".yahoo") > -1:
      return ("Yahoo Distribution")

    #
    if cDomain.find(".google") > -1:
      return ("Google Distribution")
  
    if cDomain.find(".doubleclick.net") > -1:
      return ("Google Distribution")

    if cDomain.find(".firebaseapp.com") > -1:
      return ("Google Firebase Hosting")

    #
    if cDomain.find(".instacontent.net") > -1:
      return ("Mirror Image")

    if cDomain.find(".cachefly.net") > -1:
      return ("Cachefly")

    if cDomain.find(".nyucd.net") > -1:
      return ("Coral Cache")

    if cDomain.find(".netdna-cdn.com") > -1:
      return ("MaxCDN")

    #KR 
    if cDomain.find(".navercdn.com") > -1:
      return ("NHN Corp.")
  
    if cDomain.find(".nheos.com") > -1:
      return ("NHN Corp.")

    if cDomain.find(".gscdn.net") > -1:
     return ("GS Neotek")
   
    if cDomain.find(".gscdn.com") > -1:
      return ("GS Neotek")

    if cDomain.find(".x-cdn.com") > -1:
      return ("LG U+")
  
    if cDomain.find(".cdn.cloudn.co.kr") > -1:
      return ("LG U+ Cloud N")

    if cDomain.find(".ktics.co.kr") > -1:
      return ("KT - SolutionBox")

    if cDomain.find(".hscdn.com") > -1:
      return ("Hyosung CDN")

    if cDomain.find(".cbricdns.com") > -1:
      return ("Hosted by Penta Security Systems")

    #
    if cDomain.find(".nciashield.org") > -1:
      return ("NCIA Shield Service - NATIONAL COMPUTING AND INFORMATION SERVICE")

    #
    if cDomain.find(".rhcloud.com") > -1:
      return ("Hosted by OpenShift - Red Hat's Cloud Computing Platform")

    #
    if cDomain.find(".core.pw") > -1:
      return ("Hosted by G-Core Labs S.A.")

    #
    if cDomain.find(".squarespace.com") > -1:
      return ("Hosted by SQUARESPACE")

    #
    if cDomain.find(".kinxcdn.com") > -1:
      return ("KINX CDN")

    #
    if cDomain.find(".whoisweb.net") > -1:
      return ("Hosted by Whois Web")

def find_Cname(domain):
    try:
        for rdata in dns.resolver.query(domain.strip(), 'CNAME'):
            cDomain = str(rdata.target)
            sys.stdout.write(CYAN)
            print ("CNAME is detected from {0} : {1}, CDN vendor is {2}".format(domain,cDomain,find_Vendor(cDomain)))
            sys.stdout.write(RESET)
            find_Cname(cDomain[0:len(cDomain)-1])
    except Exception:
        find_ARecord(domain)

def find_ARecord(domain):
	try:
		for rdata in dns.resolver.query(domain, 'A'):
			print("CNAME is not detected from {0}, IP Address is {1}".format(domain,rdata.address))
	except:
		print("Can't find DNS lookup: " + domain)
	finally:
		print("------------------------------------------------------------------------------------------------------------------------")

colC = ws['E']
for cell in colC:
	if (len(str(cell.value))) > 7:
		find_Cname(cell.value)


