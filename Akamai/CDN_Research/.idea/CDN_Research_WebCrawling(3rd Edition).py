import os
import xlsxwriter
import openpyxl
import urllib.request
import socket
from socket import timeout
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import re
import ssl
import dns.resolver

# 함수 정의 부문 #

def changeDir():
    os.chdir("C:/Users/yjung/Desktop/Personal/Github/AKAMAI/CDN_Research")

# http://www.xxx.com 형식으로 만들어주는 함수
def full_Url(url):
    # URL이 없다면..
    if(url==None):
        print("[ERROR] URL IS NOT PROVIDED.")
        writeExcel(worksheet, counter, 1, "[ERROR] URL UNPROVIDED")
    else:
        if ("http:" in str(url)):
            full = url
        elif ("https:" in str(url)):
            full = url
        else:
            full = "http://" + url
        print(full, "홈페이지에 접근하고 있습니다.")
        return full

# 웹사이트에서 특정 탭 가져오기
def decode(url):
    request = urllib.request.Request(url)
    #print("1 : ", request)
    temp = urllib.request.urlopen(request, timeout=5)
    #print("2 : ", temp)
    data = temp.read()
    #wwprint("3 : ", data)
    soup = BeautifulSoup(data, "lxml")

    link1 = [element.get("href") for element in soup.find_all("a")]
    link2 = [element.get("src") for element in soup.find_all("script")]
    link3 = [element.get("href") for element in soup.find_all("link")]
    link4 = [element.get("content") for element in soup.find_all("meta")]
    link5 = [element.get("value") for element in soup.find_all("option")]
    result = link1 + link2 + link3 + link4 + link5

    return deleteSameDomain(url, trimUrl(result))

# 같은 URL을 제거해주는 함수
def deleteSameDomain(original, list):
    result = []
    # print(list)

    if(list):
        for element in list:
            if (original == element):
                continue
            if(element=="://"):
                continue
            if (element in original):
                continue
            else:
                result.append(element)
    return result

# 이 키워드가 들어간 도메인은 제거.
blacklist = ["javascript",
             "facebook", "twitter", "instagram", "youtube", "youtu.be",
             "goo.gl", "google", "naver", "tistory", "amazon", "aws", "amazonaws"
             "ad.about", "mailto",
             "linkedin", "slideshare", "github", "flickr", "pinterest",
             "cloudfront", "wordpress", "cisco", "citrix", "netapp"
             "hp.com", "microsoft", "apple", "vimeo", "surveymonkey", "maxcdn",
             "jquery", "symantec", "norton", "cdnetworks", "kakao."
             "ad."]

# 특정 키워드를 포함하는 URL을 제거해주는 함수
def trimUrl (list):
    result = []
    for element in list:
        if(element == None):
            continue
        elif ("http" not in element):
            continue
        elif(element=="://"):
            continue
        for keyword in blacklist:
            if(keyword in element):
                break
        else:
            result.append(element)

    # print("3 : ", result)
    return removeExtra(result)

# 크롤링된 주소에서 불필요한 부분을 제거해주는 함수
def removeExtra(list):
    result = []
    i=1
    for element in list:
        # 약간의 노가다성 및 특별성이 있기는 하지만, 예외적인 사이트 html 코드를 타겟.
        if (";" in element):
            for element in list:
                if(" url=" in element):
                    url = re.search("(?P<url>https?://[^\s]+)", element).group("url")
                    if("lotte" in url):
                        result.append(url)
                    else:
                        removeExtra([url])
                if("action-uri=" in element):
                    url = re.search("(?P<url>https?://[^\s]+)", element).group("url")
                    removeExtra([url])

        else:
            url = urlparse(element)
            if(len(url)>1):
                add = url[0] + "://" + url[1]
                if(add in result):
                    continue
                else:
                    result.append(add)
                    # print("{} 가 추가되었습니다.".format(add))
            i += 1
    return removeSame(result)

# 동일한 결과값을 제거해주는 함수
def removeSame(list):
    result = []
    for i in range(len(list)):
        if(list[i] in result):
            continue
        else:
            result.append(list[i])
    return result

# 엑셀에 작성하는 함수
def writeExcel(file, row, column, content):
    file.write(row, column, content)

# CDN Vendor를 찾아주는 함수
def find_Vendor(cDomain):
    category = {
        "Akamai"                        : [".globalredir.akadns.net", ".akadns.net", ".edgekey.net", ".edgesuite.net", ".akamaiedge.net",".akamai.net"],
        "Level3"                        : [".nstatic.net", ".footprint.net"],
        "Microsoft"                     : [".azurewebsites.net", ".cloudapp.net", ".msecnd.net"],
        "Amazon"                        : [".elb.amazonaws.com", ".amazonaws.com", ".cloudfront.net"],
        "Wix"                           : [".wixdns.net", ".wix.com"],
        "OkServer Internet Solution Inc": [".okserver.cn", "okserver.net"],
        "Cdnetworks"                    : [".gccdn.net", ".cdnetworks.net", ".cdngc.net", ".cdngs.net", ".speedcdn.net", ".cdngl.net", ".cdnga.net"],
        "Limelight Networks"            : [".llnwd.net", ".lldns.net"],
        "Verizon - EdgeCast"            : [".v0cdn.net", ".edgecastcdn.net", ".cedexis.net", ".mucdn.net"],
        "Fastly"                        : [".fastly.net"],
        "CloudFlare"                    : [".cloudflare.net"],
        "Highwinds"                     : [".hwcdn.net"],
        "CDN77(onApp)"                  : [".cdn77.org"],
        "ChinaCache"                    : [".lxsvc.cn", ".ccgslb.com", ".ccna.c3cdn.net"],
        "Yahoo"                         : [".yahoo"],
        "Google"                        : [".google.com", ".doubleclick.net", ".firebaseapp.com"],
        "Mirror Image"                  : [".instacontent.net"],
        "Cachefly"                      : [".cachefly.net"],
        "Coral Cache"                   : [".nyucd.net"],
        "MaxCDN"                        : [".netdna-cdn.com"],
        "NHN Corp."                     : [".navercdn.com", ".nheos.com"],
        "GS Neotek"                     : [".gscdn.net", ".gscdn.com"],
        "LG"                            : [".xcdn.com", ".cdn.cloudn.co.kr"],
        "KT"                            : [".ktics.co.kr"],
        "Hyosung"                       : [".hscdn.com"],
        "Penta Security Systems"        : [".cbricdns.com"],
        "NCIA Sheild Service"           : [".nciashield.org"],
        "Redhat"                        : [".rhcloud.com"],
        "G-Core Labs S.A."              : [".core.pw"],
        "SQUARESPACE"                   : [".squarespace.com"],
        "KINX"                          : [".kinxcdn.com"],
        "Whois Web"                     : [".whoisweb.net"]
    }
    cdn_vendor = list(category.keys())

    for i in range(len(cdn_vendor)):
        for element in category.get(cdn_vendor[i]):
            if(element in cDomain):
                return cdn_vendor[i]

def end():
    input("End of Program. Enter any key to exit.....")

# 메인 함수 시작 부분
changeDir()

if(__name__ == "__main__"):

    # 데이터를 읽어오는 원본 파일
    source = openpyxl.load_workbook("CAT_List_Akamai.xlsx")
    sheet_load = source.get_sheet_by_name("CAT-4")

    # 데이터를 저장하는 대상 파일
    destination = xlsxwriter.Workbook("CDN_Research_Results.xlsx")
    worksheet = destination.add_worksheet("CAT-4")

    writeExcel(worksheet, 0, 0, "Sequence")
    writeExcel(worksheet, 0, 1, "Website")
    writeExcel(worksheet, 0, 2, "CDN Services")
    writeExcel(worksheet, 0, 3, "Name")
    writeExcel(worksheet, 0, 4, "Phone")
    writeExcel(worksheet, 0, 5, "Email")

    for row_no in range(2, sheet_load.max_row+1):
        print("현재 {} 번째 웹사이트 작업중입니다.".format(row_no - 1))

        try:
            # 방문할 홈페이지 주소를 Original 엑셀 파일에서 가져옴
            url = sheet_load.cell(row=counter, column=3).value

            # 엑셀 안에 URL이 없는 경우
            if (url == None):
                print("에러 : 엑셀 안에 URL이 없습니다.")
                print("")
                writeExcel(worksheet, rownum, 0, industry)
                writeExcel(worksheet, rownum, 1, account)
                writeExcel(worksheet, rownum, 2, "X")
                writeExcel(worksheet, rownum, 3, "-")
                writeExcel(worksheet, rownum, 4, "CAT-4")
                writeExcel(worksheet, rownum, 5, "[ERROR] URL UNPROVIDED")

            # 엑셀 안에 URL이 있는 경우
            else:

                url = full_Url(sheet_load.cell(row=counter, column=3).value)
                result = decode(url)
                writeExcel(worksheet, rownum, 0, industry)
                writeExcel(worksheet, rownum, 1, account)
                writeExcel(worksheet, rownum, 2, url)
                writeExcel(worksheet, rownum, 4, "CAT-4")

                # 메인 홈페이지 주소에서 CNAME 검색
                strip = urlparse(url)
                find_Cname_main(strip[1])
                # 크롤된 결과값이 NULL인 경우
                if (result == None):
                    print("[ERROR] NO URL IS FOUND FROM THE WEBSITE")
                    print("")
                    # print("②", end=" ")
                    writeExcel(worksheet, rownum, 5, "[ERROR] NO URL FOUND")

                # [안전장치] 크롤된 결과값이 빈 리스트인 경우
                elif (result == []):
                    print("[ERROR] NO URL IS FOUND FROM THE WEBSITE")
                    print("")
                    # print("④", end=" ")
                    writeExcel(worksheet, rownum, 5, "[ERROR] NO URL FOUND")

                # 크롤된 결과값이 뭐라도 있는 경우
                else:
                    # i는 패밀리 사이트, 해외사이트를 적기 시작할 5번째 열부터 시작
                    # j는 shell 창에서 출력할 때 번호 매김용, 1부터 시작
                    i = 4
                    j = 1
                    k = 6
                    # 최종 결과값을 리스트 형식으로 받아와서 traversing
                    for element in result:
                        # urlparse 함수를 사용하여 쓸데없는 부분을 버린다.
                        temp = urlparse(element)
                        # [요청] http:// 및 https:// 제거
                        print("{:3} {}".format(j, temp[1]))
                        result = temp[1]

                        # CNAME을 검색
                        find_Cname_sub(result)
                        writeExcel(worksheet, rownum, k, result)
                        print("")
                        i += 1
                        j += 1
                        k += 2

                    print("")
        except :
            print("!!")


    source.close()
    destination.close()
    end()